import os
import json
import time
import allure
import logging
import openpyxl
import subprocess
from collections import deque
from datetime import datetime
from gz_public import time_difference

logger = logging.getLogger(__name__)


def get_dev_udid_through_pymobiledevice3(device_name, connect_type):
    """
    :param device_name: 设备名称，例如，iPhonexxx
    :param connect_type:连接类型，例如，USB
    :return:获取对应设备的udid
    """
    try:
        result = subprocess.run(['pymobiledevice3', 'usbmux', 'list'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        if result.returncode == 0:
            devices = json.loads(result.stdout)
            device_udid = None
            for device in devices:
                if device['DeviceName'] == device_name and device['ConnectionType'] == connect_type:
                    device_udid = device['UniqueDeviceID']
                    break
            return device_udid
    except Exception as e:
        print(f"出错了:{e}")
    return None


def get_dev_udid():
    try:
        result = subprocess.run(['ideviceinfo'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        if result.returncode == 0:
            device_info = {}
            for line in result.stdout.splitlines():
                if 'DeviceName' in line:
                    device_info['DeviceName'] = line.split(":")[1].strip()
                if 'UniqueDeviceID' in line:
                    device_info['UDID'] = line.split(":")[1].strip()
            return device_info
        else:
            print(f"无法获取设备信息：{result.stderr}")
    except Exception as e:
        print(f"出错了{e}")
    return None


def get_dev_play_state_through_libimobiledevice(iphone_model, sn, log_date):
    """
    :param iphone_model: iPhone型号，例如，iPhoneX，本地根据手机型号创建挂载目录
    :param sn: 设备sn
    :param state: 不同开流状态，例如，wakeStart、awake success、p2pStart、p2pEnd、previewStart、previewSuccess
    :param log_date: 日志日期，例如，20241028
    :return:
    """
    filter_condition = f"'wakeStart  deviceId = {sn}'"
    mount_path = f"/Users/testmanzhang/ios_sandbox/{iphone_model}/Documents/Logs/"
    log_file = f"glazero_app_ios_{log_date}.log"

    log_file_path = f"{mount_path}{log_file}"

    # 刷新挂载目录
    refresh_cmd = f"ls {mount_path}"
    result_ls = subprocess.run(refresh_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    # print("Directory contents:\n", result_ls.stdout)
    if result_ls.stderr:
        print("刷新挂载目录Error:\n", result_ls.stderr)

    # 从日志文件中获取状态
    result_get = None
    get_state_cmd = f"grep {filter_condition} {mount_path}{log_file}"
    if os.path.isfile(log_file_path):
        result_get = subprocess.run(get_state_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        # print(result_get.stdout)
        if result_get.stderr:
            print("grep 错误:", result_get.stderr)
    else:
        print("日志文件不存在:", log_file_path)

    # 分析获取的结果
    if result_get:
        lines = result_get.stdout.splitlines()
        last_success_line = lines[-1]
        print("最后一行唤醒成功的结果：\n", last_success_line)
        results_split = last_success_line.split()

        # 计算唤醒成功的时间，收到唤醒成功结果的时间点-发送唤醒指令的时间点


# iOS开流过程中无此状态
def get_start_create_LivePlay_fragment_result(dev_id, click_time):
    """
    目前iOS没有这个状态，所以不统计这个节点
    :param dev_id:
    :param click_time:
    :return:
    """
    # 获取 APP 从首页进入开流页面时  的日志
    # C9S  LivePlayFragment:onCreate改成了:LivePlaySingleFragment:onCreate
    cmd = 'adb -s %s shell logcat -v time -s BaseFragment "\| grep -e LivePlaySingleFragment:onCreate"' % dev_id
    print(cmd)

    p_obj = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 等待日志获取完成
    time.sleep(4)
    # 获取完成后药结束子进程
    p_obj.terminate()
    p_obj.kill()
    # 获取结束后读取内容
    lines = p_obj.stdout.readlines()
    print(lines)
    # 如果获取到的内容为空执行返回值状态为空,否则,根据返回内容,返回成功或失败
    if len(lines) == 0:
        create_LivePlay_fragment_state = 'create_fragment:null'
        print('开流播放片段在创建状态为空')
        # 将从点击屏幕 到 进入开流页面 的次数,存储在 ./data.xlsx 路径文件中的 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
        colume_list_to_add_null = ['AJ', 'AK', 'AL', 'AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_list_to_add_null)
        # 将从点击屏幕 到 进入开流页面 的时间,存储在 ./data.xlsx 路径文件中的 'B', 'C', 'D', 'E', 'F', 'G' 列最后一行
        colume_to_add_null = ['B', 'C', 'D', 'E', 'F', 'G']
        content_to_add_null = ['进入开流页面为空', '开始唤醒为空', '唤醒为空', 'p2p连接为空', 'Preview 为空',
                               '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)

        # 获取完成后清除logcat
        r_obj = os.popen("adb logcat -c")
        r_obj.close()
    else:
        # 将bytes转换成字符串
        last_line = lines[-2].decode('utf-8')  # -1 改成 -2
        print("最后一行：", last_line)
        # 计算 从点击屏幕 到 进入开流页面 的时间差值
        time_difference_success = time_difference(last_line, click_time)
        # 将 last_line 的内容赋值给 results
        results = last_line.split()
        print(results)
        # 查找 onCreate 字段
        for word in results:
            if 'LivePlaySingleFragment:onCreate' in word:
                create_LivePlay_fragment_state = word
                print('开始创建开流现场播放片段(进入开流页面):', create_LivePlay_fragment_state)

                # 将从点击屏幕开流 到 进入开流页面 的次数,存储在 ./data.xlsx 路径文件中的 AJ列最后一行
                colume_to_add_success = ['AJ']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将从点击屏幕开流 到 进入开流页面 的时间,存储在 ./data.xlsx 路径文件中的 AB 列最后一行
                colume_to_add_success = ['B']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)

                break
        else:
            create_LivePlay_fragment_state = 'create_fragment:Failed'  # 如果最后一行中,没有wakeSuccess即 设备唤醒 失败, dev_wake_state 赋值为失败状态.
            print('设备未收到唤醒指令:', create_LivePlay_fragment_state)

            # 将从点击屏幕开流 到 进入开流页面 的次数,存储在 ./data.xlsx 路径文件中的 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
            colume_to_add_failed = ['AJ', 'AK', 'AL', 'AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
            # 将从点击屏幕开流 到 进入开流页面 的时间,存储在 ./data.xlsx 路径文件中的 'B', 'C', 'D', 'E', 'F', 'G' 列最后一行
            colume_to_add_failed = ['B', 'C', 'D', 'E', 'F', 'G']
            content_to_add_failed = ['进入开流页面失败', '开始唤醒失败', '唤醒失败', 'p2p连接失败', 'Preview 失败',
                                     '休眠失败']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)

            # 获取完成后清除logcat
            r_obj = os.popen("adb logcat -c")
            r_obj.close()

    return create_LivePlay_fragment_state


def get_dev_start_wake_state_result(click_time, iphone_model, sn):
    """
    :param click_time: 点击设备的时间点
    :param iphone_model: iPhone型号，例如，iPhoneX，本地根据手机型号创建挂载目录
    :param sn: 设备sn，例如，C6L2BA110004740
    :return: 返回start_wake_state，例如，wakeStart
    """
    grep_condition = f"'wakeStart  deviceId = {sn}'"
    result_get_success = grep_from_sandbox_log(grep_condition, iphone_model)

    # 分析获取的结果
    if result_get_success:
        lines = result_get_success.splitlines()
        last_line = lines[-1]

        print("最后一行开始唤醒成功的结果：\n", last_line)
        logger.info(f"最后一行开始唤醒成功的结果: {last_line}\n")

        # 计算从 点击 设备卡片到 开始唤醒 的时间差值
        time_difference_success = time_difference(last_line, click_time)

        if time_difference_success:
            # 将 last_line 的内容赋值给 results，转换成列表
            results = last_line.split()
            print(results)

            # 查找 wakeStart 字段
            if 'wakeStart' in results:
                start_wake_state = 'wakeStart'
                print('设备收到开始唤醒指令:', start_wake_state)
                logger.info(f"设备收到开始唤醒指令: {start_wake_state}")

                # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的次数,存储在 ./data.xlsx 路径文件中的 AK列最后一行
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1,
                                              ['AK'])

                # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的时间,存储在 ./data.xlsx 路径文件中的 AC 列最后一行
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success,
                                              ['C'])
            else:
                # 如果最后一行中,没有 wakeSuccess 即 设备唤醒 失败, dev_wake_state 赋值为失败状态。
                start_wake_state = 'startwakeFailed'
                print('设备未收到开始唤醒指令:', start_wake_state)
                logger.info(f"设备未收到开始唤醒指令: {start_wake_state}")

                # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的次数,存储在 ./data.xlsx 路径文件中的 'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
                colume_to_add_failed = ['AK', 'AL', 'AM', 'AN', 'AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 0,
                                              colume_to_add_failed)

                # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的时间,存储在 ./data.xlsx 路径文件中的 'C', 'D', 'E', 'F', 'G' 列最后一行
                colume_to_add_failed = ['C', 'D', 'E', 'F', 'G']
                content_to_add_failed = ['开始唤醒失败', '唤醒失败', 'p2p连接失败', 'Preview 失败', '休眠失败']
                result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed,
                                            colume_to_add_failed)

        else:
            start_wake_state = 'startwakeState:null'
            print('开始唤醒状态为空')
            logger.info('开始唤醒状态为空, startwakeState:null')

            # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的次数,存储在 ./data.xlsx 路径文件中的  'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
            colume_list_to_add_null = ['AK', 'AL', 'AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0,
                                          colume_list_to_add_null)

            # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的时间,存储在 ./data.xlsx 路径文件中的  'C', 'D', 'E', 'F', 'G' 列最后一行
            colume_to_add_null = ['C', 'D', 'E', 'F', 'G']
            content_to_add_null = ['开始唤醒为空', '唤醒为空', 'p2p连接为空', 'Preview 为空', '休眠为空']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null,
                                        colume_to_add_null)
    else:
        start_wake_state = 'startwakeState:null'
        print('iOS沙盒目录中的日志文件不存在，有可能是0点，日志文件没有生成，延迟了1-1.5分钟')
        logger.info('iOS沙盒目录中的日志文件不存在，有可能是0点，日志文件没有生成，延迟了1-1.5分钟')
        logger.info('开始唤醒状态为空, startwakeState:null，基于上面的原因要在excel中写入空值，不然后面的执行会占用这个excel格的位置')

        # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的次数,存储在 ./data.xlsx 路径文件中的  'AK', 'AL', 'AM', 'AN', 'AO' 列最后一行
        colume_list_to_add_null = ['AK', 'AL', 'AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0,
                                      colume_list_to_add_null)

        # 将从 点击屏幕开流 到 设备收到 开始唤醒时 的时间,存储在 ./data.xlsx 路径文件中的  'C', 'D', 'E', 'F', 'G' 列最后一行
        colume_to_add_null = ['C', 'D', 'E', 'F', 'G']
        content_to_add_null = ['开始唤醒为空', '唤醒为空', 'p2p连接为空', 'Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null,
                                    colume_to_add_null)

    return start_wake_state


def get_dev_wake_state_result(click_time, iphone_model, sn):
    grep_condition = f"'awake success, deviceId={sn}'"
    result_get_success = grep_from_sandbox_log(grep_condition, iphone_model)

    # 分析获取的结果
    if result_get_success:
        lines = result_get_success.splitlines()
        last_line = lines[-1]
        print("最后一行唤醒成功的结果：\n", last_line)
        logger.info(f"最后一行唤醒成功的结果: {last_line}\n")

        # 计算从 点击 设备卡片到 开始唤醒 的时间差值
        time_difference_success = time_difference(last_line, click_time)

        if time_difference_success:
            # 将 last_line 的内容赋值给 results，转换成列表
            results = last_line.split()
            print(results)
            logger.info(f"切割后的结果是: {results}")

            find_awake_success = any(a == 'awake' and b == 'success,' for a, b in zip(results, results[1:]))

            if find_awake_success:
                dev_wake_state = 'awake success'
                print("唤醒成功: ", dev_wake_state)
                logger.info(f"唤醒成功: {dev_wake_state}")

                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 AL列最后一行
                colume_to_add_success = ['AL']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 AD 列最后一行
                colume_to_add_success = ['D']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
            else:
                # 如果最后一行中,没有wakeSuccess即 设备唤醒 失败, dev_wake_state 赋值为失败状态。
                dev_wake_state = 'wakeFailed'
                print('唤醒失败:', dev_wake_state)
                logger.error(f"唤醒失败: {dev_wake_state}")
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 'AL', 'AM', 'AN', 'AO' 列最后一行
                colume_to_add_failed = ['AL', 'AM', 'AN', 'AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 B、C、D、E 列最后一行
                colume_to_add_failed = ['D', 'E', 'F', 'G']
                content_to_add_failed = ['唤醒失败', 'p2p连接失败', 'Preview 失败', '休眠失败']
                result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)
        else:
            dev_wake_state = 'wakeState:null'
            print('唤醒状态为空', dev_wake_state)
            logger.info(f"唤醒状态为空, {dev_wake_state}")
            # 将从点击屏幕开流到设备唤醒成功的次数,存储在 ./data.xlsx 路径文件中的 'AL', 'AM', 'AN', 'AO' 列最后一行
            colume_list_to_add_null = ['AL', 'AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_list_to_add_null)
            # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 B、C、D、E 列最后一行
            colume_to_add_null = ['D', 'E', 'F', 'G']
            content_to_add_null = ['唤醒为空', 'p2p连接为空', 'Preview 为空', '休眠为空']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)
    else:
        dev_wake_state = 'wakeState:null'
        logger.info('iOS沙盒目录中的日志文件不存在，有可能是0点，日志文件没有生成，延迟了1-1.5分钟')
        logger.info('唤醒状态为空, wakeState:null，基于上面的原因要在excel中写入空值，不然后面的执行会占用这个excel格的位置')
        print('唤醒状态为空', dev_wake_state)
        logger.info(f"唤醒状态为空, {dev_wake_state}")
        # 将从点击屏幕开流到设备唤醒成功的次数,存储在 ./data.xlsx 路径文件中的 'AL', 'AM', 'AN', 'AO' 列最后一行
        colume_list_to_add_null = ['AL', 'AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_list_to_add_null)
        # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 B、C、D、E 列最后一行
        colume_to_add_null = ['D', 'E', 'F', 'G']
        content_to_add_null = ['唤醒为空', 'p2p连接为空', 'Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)

    return dev_wake_state


def result_save_excel(file_path, Sheet, content_to_add, column_to_add):
    # 打开Excel文件
    file_path = file_path
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[Sheet]
    # 要添加的内容
    content_to_add = content_to_add
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    column_to_add = column_to_add
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    for row in range(max_row, 0, -1):
        # 获取 想要添加的列中 行数内的内容.
        cell_value = sheet[f'{column_to_add}{row}'].value
        # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
        if cell_value is not None:
            last_row = row
            break
    if last_row is not None:
        # 在列的最后一个非空单元格的下一行追加内容
        sheet[f'{column_to_add}{last_row + 1}'] = content_to_add

    # 保存修改后的Excel文件
    workbook.save(file_path)
    # 关闭 Excel 文件
    workbook.close()


def result_save_excel_column_list(file_path, sheet, content_to_add, column_list_to_add):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[sheet]
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    for item in column_list_to_add:
        column = item
        for row in range(max_row, 0, -1):
            # 获取 想要添加的列中 行数内的内容.
            cell_value = sheet[f'{column}{row}'].value
            # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
            if cell_value is not None:
                last_row = row
                break
        if last_row is not None:
            # 在列的最后一个非空单元格的下一行追加内容
            sheet[f'{column}{last_row + 1}'] = content_to_add
    # 保存修改后的Excel文件
    workbook.save(file_path)
    # 关闭 Excel 文件
    workbook.close()


def result_save_excel_full_list(file_path, sheet, content_full_list_to_add, column_full_list_to_add):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表（目前是Sheet1,默认Sheet1,注意大小写,如果表格中工作表名称是 Sheet1 ,写成小写的 sheet 代码会执行失败）
    sheet = workbook[sheet]
    # 获取列的最大行号
    max_row = sheet.max_row
    # 找到列中的最后一个非空单元格
    last_row = None
    # 要写入的列, ***** 注意:写入列column_to_add的类型为列表list !!!!!
    # 要添加的内容
    # 如果所在列 行内,内容不为空,继续执行,直到行内容为空,跳出循环
    i = 0
    for item in column_full_list_to_add:
        column = item
        for row in range(max_row, 0, -1):
            # 获取 想要添加的列中 行数内的内容.
            cell_value = sheet[f'{column}{row}'].value
            # 如果所在列 行内,内容不为空, 将当前行数赋值给 loast_row 继续执行,直到行内容为空,即当前最后一行有数据的行数。跳出循环.
            if cell_value is not None:
                last_row = row
                break
        if last_row is not None:
            # 在列的最后一个非空单元格的下一行追加内容
            content = content_full_list_to_add[i]
            sheet[f'{column}{last_row + 1}'] = content
        i = i + 1
    # 保存修改后的Excel文件
    workbook.save(file_path)

    # 关闭 Excel 文件
    workbook.close()


def get_dev_p2p_state_result(click_time, iphone_model, sn):
    grep_condition = f"'connectP2P.* {sn}connect Succ'"
    result_get_success = grep_from_sandbox_log(grep_condition, iphone_model)

    # 分析获取的结果
    if result_get_success:
        lines = result_get_success.splitlines()
        last_line = lines[-1]
        print("最后一行p2p连接成功的结果：\n", last_line)
        logger.info(f"最后一行p2p连接成功的结果: {last_line}\n")

        # 计算从 点击 设备卡片到 开始唤醒 的时间差值
        time_difference_success = time_difference(last_line, click_time)

        if time_difference_success:
            # 将 last_line 的内容赋值给 results，转换成列表
            results = last_line.split()
            print(results)
            logger.info(f"切割后的结果是: {results}")
            elements_to_check = ['connectP2P', f'{sn}connect', 'Succ']

            find_p2p_connect_success = all(element in results for element in elements_to_check)

            if find_p2p_connect_success:
                dev_p2p_state = f'connectP2P {sn}connect Succ'
                print("p2p连接成功: ", dev_p2p_state)
                logger.info(f"p2p连接成功: {dev_p2p_state}")

                # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 AM 列最后一行
                colume_to_add_success = ['AM']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 E  列最后一行
                colume_to_add_success = ['E']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
            else:
                # 如果最后一行中,没有success即p2p连接为失败, dev_p2p_state 赋值为失败状态。
                dev_p2p_state = 'p2pState:failed'
                print(f"p2p连接失败：{dev_p2p_state}")
                logger.info(f"p2p连接失败：{dev_p2p_state}")
                # 将 从点击屏幕开流 到 p2p唤醒成功 的次数,存储在 ./data.xlsx 路径文件中的 'AM', 'AN', 'AO' 列最后一行
                colume_to_add_failed = ['AM', 'AN', 'AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)

                # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 'E', 'F', 'G' 列最后一行
                colume_to_add_failed = ['E', 'F', 'G']
                content_to_add_failed = ['p2p连接失败', 'Preview 失败', '休眠失败']
                result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)

        else:
            dev_p2p_state = 'p2pState:null'
            print(f"p2p state获取失败：{dev_p2p_state}")
            logger.info(f"p2p state获取失败：{dev_p2p_state}")
            # 将 从点击屏幕开流 到 p2p唤醒成功 的次数,存储在 ./data.xlsx 路径文件中的 'AM', 'AN', 'AO' 列最后一行
            colume_to_add_null = ['AM', 'AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
            # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 'E', 'F', 'G' 列最后一行
            colume_to_add_null = ['E', 'F', 'G']
            content_to_add_null = ['p2p连接为空', 'Preview 为空', '休眠为空']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)
    else:
        dev_p2p_state = 'p2pState:null'
        logger.info('iOS沙盒目录中的日志文件不存在，有可能是0点，日志文件没有生成，延迟了1-1.5分钟')
        logger.info('p2p连接状态为空, p2pState:null，基于上面的原因要在excel中写入空值，不然后面的执行会占用这个excel格的位置')
        print(f"p2p state获取失败：{dev_p2p_state}")
        logger.info(f"p2p state获取失败：{dev_p2p_state}")
        # 将 从点击屏幕开流 到 p2p唤醒成功 的次数,存储在 ./data.xlsx 路径文件中的 'AM', 'AN', 'AO' 列最后一行
        colume_to_add_null = ['AM', 'AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将 从点击屏幕开流 到 p2p唤醒成功 的时间,存储在 ./data.xlsx 路径文件中的 'E', 'F', 'G' 列最后一行
        colume_to_add_null = ['E', 'F', 'G']
        content_to_add_null = ['p2p连接为空', 'Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)

    return dev_p2p_state


def get_dev_play_state_result(click_time, iphone_model, sn):
    grep_condition = f"'previewSuccess  deviceId = {sn}'"
    result_get_success = grep_from_sandbox_log(grep_condition, iphone_model)

    # 分析获取的结果
    if result_get_success:
        logger.info(r"获取到的grep结果非空，按照换行符 \n 将日志拆分成一个列表")
        lines = result_get_success.splitlines()
        last_line = lines[-1]
        print("最后一行preview成功的结果：\n", last_line)
        logger.info(f"最后一行preview成功的结果: {last_line}\n")

        # 计算从 点击 设备卡片到 开始唤醒 的时间差值
        time_difference_success = time_difference(last_line, click_time)

        if time_difference_success:
            # 将 last_line 的内容赋值给 results，转换成列表
            results = last_line.split()
            print(results)
            logger.info(f"切割后的结果是: {results}")
            elements_to_check = ['previewSuccess', f'{sn}']

            find_preview_success = all(element in results for element in elements_to_check)

            if find_preview_success:
                dev_preview_state = 'previewSuccess'
                print("preview成功: ", dev_preview_state)
                logger.info(f"preview成功: {dev_preview_state}")

                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 AN 列最后一行
                colume_to_add_success = ['AN']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将从点击屏幕开流到设备唤醒成功的时间,存储在 ./data.xlsx 路径文件中的 F 列最后一行
                colume_to_add_success = ['F']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
            else:
                # 如果最后一行中,没有 Playing 即 第一帧回调失败, play_state 赋值为失败状态。
                dev_preview_state = 'state:failed'
                # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'AN', 'AO' 列最后一行
                colume_to_add_failed = ['AN', 'AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
                # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'F', 'G' 列最后一行
                colume_to_add_failed = ['F', 'G']
                content_to_add_failed = ['Preview 失败', '休眠失败']
                result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_failed, colume_to_add_failed)

        else:
            # 如果最后一行中,没有 Playing 即 第一帧回调失败, play_state 赋值为失败状态。
            dev_preview_state = 'state:null'
            # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'AN', 'AO' 列最后一行
            colume_to_add_null = ['AN', 'AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
            # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'F', 'G' 列最后一行
            colume_to_add_null = ['F', 'G']
            content_to_add_null = ['Preview 为空', '休眠为空']
            result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)
    else:
        # 如果最后一行中,没有 Playing 即 第一帧回调失败, play_state 赋值为失败状态。
        dev_preview_state = 'state:null'
        logger.info('iOS沙盒目录中的日志文件不存在，有可能是0点，日志文件没有生成，延迟了1-1.5分钟')
        logger.info('preview状态为空, state:null，基于上面的原因要在excel中写入空值，不然后面的执行会占用这个excel格的位置')
        print(f"preview state获取失败：{dev_preview_state}")
        logger.info(f"preview state获取失败：{dev_preview_state}")
        # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'AN', 'AO' 列最后一行
        colume_to_add_null = ['AN', 'AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将从点击屏幕开流到设备第一帧画面成功回来的时间,存储在 ./data.xlsx 路径文件中的 'F', 'G' 列最后一行
        colume_to_add_null = ['F', 'G']
        content_to_add_null = ['Preview 为空', '休眠为空']
        result_save_excel_full_list('./data.xlsx', 'Sheet1', content_to_add_null, colume_to_add_null)

    return dev_preview_state


def get_dev_dormancy_state_result(click_time, iphone_model, sn):
    grep_condition = f"'\"deviceId\":\"{sn}\",\"dp\":\"149\",\"value\":\"0\"'"
    result_get_success = grep_from_sandbox_log(grep_condition, iphone_model)

    # 分析获取的结果
    if result_get_success:
        lines = result_get_success.splitlines()
        last_line = lines[-1]
        print("最后一行休眠成功的结果：\n", last_line)
        logger.info(f"最后一行休眠成功的结果: {last_line}\n")

        # 计算从 点击 设备卡片到 开始唤醒 的时间差值
        time_difference_success = time_difference(last_line, click_time)

        if time_difference_success:
            # 将 last_line 的内容赋值给 results，转换成列表
            results = last_line.split()
            print(results)
            logger.info(f"切割后的结果是: {results}")

            if f'{{"deviceId":"{sn}","dp":"149","value":"0"}}' in results:
                dev_dormancy_state = f'{{"deviceId":"{sn}","dp":"149","value":"0"}}'
                print("休眠成功: ", dev_dormancy_state)
                logger.info(f"休眠成功: {dev_dormancy_state}")

                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
                colume_to_add_success = ['AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 1, colume_to_add_success)
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
                colume_to_add_success = ['G']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', time_difference_success, colume_to_add_success)
            else:
                # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态。
                dev_dormancy_state = f'{{"deviceId":"{sn}","dp":"149","value":"1"}}'
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
                colume_to_add_failed = ['AO']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_failed)
                # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
                colume_to_add_failed = ['G']
                result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠失败', colume_to_add_failed)

        else:
            # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态。
            dev_dormancy_state = f'{{"deviceId":"{sn}","dp":"149","value":Null}}'
            # 将 从点击屏幕返回 到 设备休眠成功 的次数,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
            colume_to_add_null = ['AO']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
            # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
            colume_to_add_null = ['G']
            result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠状态为空', colume_to_add_null)
    else:
        # 如果最后一行中,没有 '\"149\":false' 即 设备依然在唤醒状态。
        dev_dormancy_state = f'{{"deviceId":"{sn}","dp":"149","value":Null}}'
        logger.info('iOS沙盒目录中的日志文件不存在，有可能是0点，日志文件没有生成，延迟了1-1.5分钟')
        logger.info('休眠状态为空, dormancy state:null，基于上面的原因要在excel中写入空值，不然后面的执行会占用这个excel格的位置')
        print(f"dormancy state获取失败：{dev_dormancy_state}")
        logger.info(f"dormancy state获取失败：{dev_dormancy_state}")
        # 将 从点击屏幕返回 到 设备休眠成功 的次数,存储在 ./data.xlsx 路径文件中的 'AO' 列最后一行
        colume_to_add_null = ['AO']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', 0, colume_to_add_null)
        # 将 从点击屏幕返回 到 设备休眠成功 的时间,存储在 ./data.xlsx 路径文件中的 G 列最后一行
        colume_to_add_null = ['G']
        result_save_excel_column_list('./data.xlsx', 'Sheet1', '休眠状态为空', colume_to_add_null)

    return dev_dormancy_state


def grep_from_sandbox_log(grep_condition, iphone_model):
    # iOS沙盒目录对应的本地挂载目录；当前的操作会在当前日期的app日志里出现
    log_date = datetime.now().strftime("%Y%m%d")
    mount_path = f"/Users/testmanzhang/ios_sandbox/{iphone_model}/Documents/Logs/"
    log_file_path = f"{mount_path}glazero_app_ios_{log_date}.log"

    # 刷新挂载目录
    refresh_cmd = f"ls {mount_path}"
    result_ls = subprocess.run(refresh_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if result_ls.stderr:
        print("刷新挂载目录Error:\n", result_ls.stderr)
        logger.error(f"刷新挂载目录Error:\n {result_ls.stderr}")

    # 从日志文件中获取状态
    get_result = None

    get_state_cmd = f"grep {grep_condition} {log_file_path}"
    logger.info(f"执行的grep命令是：{get_state_cmd}")

    if os.path.isfile(log_file_path):
        result = subprocess.run(get_state_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

        if result.stderr:
            logger.error(f"grep 错误: \n {result.stderr}")
        elif result.returncode == 0 and result.stdout == '':
            logger.info(f"返回的return code为''，stdout为''")
        elif result.returncode == 0 and result.stdout != '':
            logger.info(f"返回的内容不为空，返回result.stdout的内容")
            get_result = result.stdout
    else:
        logger.info(f"日志文件不存在: {log_file_path}")

    return get_result


def clear_sandbox_log(iphone_model):
    # iOS沙盒目录对应的本地挂载目录；当前的操作会在当前日期的app日志里出现
    log_date = datetime.now().strftime("%Y%m%d")
    mount_path = f"/Users/testmanzhang/ios_sandbox/{iphone_model}/Documents/Logs/"
    log_file_path = f"{mount_path}glazero_app_ios_{log_date}.log"

    # 刷新挂载目录
    refresh_cmd = f"ls {mount_path}"
    result_ls = subprocess.run(refresh_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if result_ls.stderr:
        print("刷新挂载目录Error:\n", result_ls.stderr)
        logger.error(f"刷新挂载目录Error:\n {result_ls.stderr}")

    # 从日志文件中获取状态
    result = None

    clear_cmd = f"truncate -s 0 {log_file_path}"
    print(f"在执行测试之前清除沙盒日志：{clear_cmd}")
    logger.info(f"在执行测试之前清除沙盒日志：{clear_cmd}")

    if os.path.isfile(log_file_path):
        result = subprocess.run(clear_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        if result.stderr:
            print("清除沙盒日志 错误:", result.stderr)
            logger.error(f"清除沙盒日志 错误: \n {result.stderr}")
        else:
            logger.info("清除完成！")
    else:
        print("日志文件不存在: ", log_file_path)
        logger.info(f"日志文件不存在: {log_file_path}")

    return result


def get_app_log(log_type, iphone_model, log_path, current_time, log_lines):
    """
    获取app日志或者涂鸦日志
    :param log_type: app 或者 ty
    :param iphone_model: 例如，iPhoneX，从而找到本地的沙盒目录
    :param log_path：例如，'./report/C6L/log_attach'
    :param log_lines：例如，1000，获取最新的1000行日志
    :param current_time：例如，5-26-31，这个要调用时传过来，因为case里面要添加附件到allure报告中
    :结果: 将从沙盒目录中app日志或者ty日志获取到的log，保存到文件，作为allure的attachment
    """
    # 获取当前日期和时间
    log_date = datetime.now().strftime("%Y%m%d")

    # iOS沙盒目录对应的本地挂载目录
    mount_path = os.path.join("/Users/testmanzhang/ios_sandbox", iphone_model, "Documents", "Logs")

    # 根据日志类型确定日志文件路径
    log_file_path = None
    if log_type == 'app':
        log_file_path = os.path.join(mount_path, f"glazero_app_ios_{log_date}.log")
    elif log_type == 'ty':
        log_file_path = os.path.join(mount_path, f"glazero_app_ios_ty_{log_date}.log")

    # log_data = get_last_lines_from_file(log_file_path, log_lines)
    log_data = get_last_lines_from_file_by_file_offset_mode(log_file_path, log_lines)

    if not log_data:
        print(f"无法取到日志数据：{log_file_path}")
        logger.error(f"无法取到日志数据：{log_file_path}")
        return

    log_file_save_path = os.path.join(log_path, f"{log_type}_log_{current_time}.log")

    with open(log_file_save_path, 'w') as file:
        for line in log_data:
            file.write(line)


def get_last_lines_from_file(file_path, num_lines):
    with open(file_path, 'r') as file:
        last_lines = deque(file, maxlen=num_lines)
    return list(last_lines)


def get_last_lines_from_file_by_file_offset_mode(file_path, num_lines):
    """
    使用文件偏移量方式获取日志
    :param file_path:日志文件路径
    :param num_lines:行数
    :return:
    """
    if not os.path.isfile(file_path):
        print(f"日志文件不存在：{file_path}")
        logger.error(f"日志文件不存在：{file_path}")
        return []

    with open(file_path, 'rb') as file:
        file.seek(0, 2)
        file_size = file.tell()

        lines_found = []
        buffer = bytearray()

        while len(lines_found) <= num_lines and file_size > 0:
            file_size -= 1
            file.seek(file_size)
            byte = file.read(1)
            buffer.extend(byte)

            if byte == b'\n':
                line = buffer[::-1].decode(errors='ignore').strip()
                lines_found.append(line + '\n')
                buffer = bytearray()

        # 当文件内容小于指定的行数是执行到这里，这行是最顶部的那一行
        if buffer:
            line = buffer[::-1].decode(errors='ignore').strip()
            lines_found.append(line)

    return list(reversed(lines_found))


def log_process_after_case_fail(driver, iphone_model, dev_model):
    # 获取当前时间，用于区分不同的日志文件作为附件。
    current_time = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())

    # 获取app日志
    get_app_log('app', iphone_model, f"./report/{dev_model}/log_attach",
                current_time, 2000)
    # 将日志添加到报告中
    allure.attach.file(f"./report/{dev_model}/log_attach/app_log_{current_time}.log",
                       name="app log", attachment_type=allure.attachment_type.TEXT)
    driver.implicitly_wait(10)

    # 获取ty日志
    # 每次重启app后，ty日志会停止打印，所以先不取ty日志
    """
    get_app_log('ty', iphone_model, f"./report/{dev_model}/log_attach",
                current_time, 4000)
    # 将日志添加到报告中
    allure.attach.file(f"./report/{dev_model}/log_attach/ty_log_{current_time}.log",
                       name="ty log", attachment_type=allure.attachment_type.TEXT)
    driver.implicitly_wait(10)
    """


# if __name__ == "__main__":
    # result_done = clear_sandbox_log_before_running('iPhoneX')
    # print(result_done)
    # hello_test_123()
    # get_dev_play_state_through_pymobiledevice3()
    # get_dev_play_state_through_libimobiledevice('iPhoneX', 'C6L2BA110004740',  "20241029")
    # click_time = "2024-11-11 10:56:36.911816"
    # time_obj = datetime.strptime(click_time, '%Y-%m-%d %H:%M:%S.%f')
    # get_dev_start_wake_state_result(time_obj, 'iPhoneX', 'C6L2BA110004740')
    # get_dev_wake_state_result(time_obj, 'iPhoneX', 'C6L2BA110004740')
    # get_dev_p2p_state_result(time_obj, 'iPhoneX', 'C6L2BA110004740')
    # get_dev_play_state_result(time_obj, 'iPhoneX', 'C6L2BA110004740')
    # get_dev_dormancy_state_result(time_obj, 'iPhoneX', 'C6L2BA110004740')
    # current_time = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())
    # get_app_log('ty', 'iPhoneX', './report/C6L/log_attach', current_time, 4000)
    # print(get_dev_udid_through_pymobiledevice3('GlazeroCom的iPhone 勿动', 'USB'))
    # print(get_device_udid_through_libimobiledevice())
    # execution_time1 = timeit.timeit(get_device_udid_through_pymobiledevice3('iPhonexxx', 'USB'), number=1)
    # execution_time2 = timeit.timeit(get_device_udid_through_libimobiledevice(), number=1)
    # print(execution_time1, execution_time2)
